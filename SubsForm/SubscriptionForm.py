import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import datetime, date
import requests
import configparser
import os
import sys
from openpyxl import load_workbook

class SubscriptionFormApp:
    def __init__(self, master):
        self.master = master
        self.search_var = tk.StringVar()
        master.title("Formulário de Assinatura")

        if getattr(sys, 'frozen', False):
            self.script_dir = os.path.dirname(sys.executable)
        else:
            self.script_dir = os.path.dirname(os.path.abspath(__file__))

        config_file_path = os.path.join(self.script_dir, "config.ini")

        if not os.path.exists(config_file_path):
            self.create_config_file(config_file_path)

        self.config = configparser.ConfigParser()
        self.config.read(config_file_path)

        self.host = self.config.get("Form", "host")
        self.port = self.config.getint("Form", "port")

        self.form_frame = ttk.Frame(master, padding="20")
        self.form_frame.grid(row=0, column=0, sticky="nsew")
        self.form_frame.columnconfigure(1, weight=1)

        ttk.Label(self.form_frame, text="Nome do Cliente:").grid(row=0, column=0, sticky="w")
        self.client_name_entry = ttk.Entry(self.form_frame)
        self.client_name_entry.grid(row=0, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Data de Término (AAAA-MM-DD):").grid(row=1, column=0, sticky="w")
        self.end_date_entry = ttk.Entry(self.form_frame)
        self.end_date_entry.grid(row=1, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Chave de Licença (opcional):").grid(row=2, column=0, sticky="w")
        self.license_key_entry = ttk.Entry(self.form_frame)
        self.license_key_entry.grid(row=2, column=1, sticky="ew", padx=(0, 10))

        ttk.Label(self.form_frame, text="Produtos:").grid(row=3, column=0, sticky="w")

        self.product_listbox = tk.Listbox(self.form_frame, selectmode=tk.SINGLE)
        self.product_listbox.grid(row=4, column=0, columnspan=2, padx=(0, 10), pady=5, sticky="nsew")
        self.product_listbox.bind("<<ListboxSelect>>", self.on_product_select)
        self.update_product_list()

        self.product_listbox.selection_clear(0, tk.END)

        ttk.Label(self.form_frame, text="Novo Produto:").grid(row=5, column=0, sticky="w")
        self.new_product_entry = ttk.Entry(self.form_frame)
        self.new_product_entry.grid(row=5, column=1, sticky="ew", padx=(0, 10))

        self.add_product_button = ttk.Button(self.form_frame, text="Adicionar Produto", command=self.add_product)
        self.add_product_button.grid(row=5, column=2, sticky="ew", padx=(0, 10))

        self.delete_product_button = ttk.Button(self.form_frame, text="Excluir Produto", command=self.delete_product)
        self.delete_product_button.grid(row=5, column=3, sticky="ew", padx=(0, 10))

        self.buttons_frame = ttk.Frame(master)
        self.buttons_frame.grid(row=1, column=0, pady=(10, 0))
        self.buttons_frame.columnconfigure((0, 1, 2, 3, 4), weight=1)

        self.add_button = ttk.Button(self.buttons_frame, text="Adicionar Assinatura", command=self.add_subscription)
        self.add_button.grid(row=0, column=0, pady=5, padx=(0, 5), sticky="ew")

        self.view_button = ttk.Button(self.buttons_frame, text="Visualizar Assinaturas", command=self.view_subscriptions)
        self.view_button.grid(row=0, column=1, pady=5, padx=(0, 5), sticky="ew")

        self.delete_button = ttk.Button(self.buttons_frame, text="Excluir Assinatura", command=self.delete_subscription)
        self.delete_button.grid(row=0, column=2, pady=5, padx=(0, 5), sticky="ew")

        self.renew_button = ttk.Button(self.buttons_frame, text="Renovar Assinatura", command=self.renew_subscription)
        self.renew_button.grid(row=0, column=3, pady=5, padx=(0, 5), sticky="ew")

        self.import_button = ttk.Button(self.buttons_frame, text="Importar do Excel", command=self.import_from_excel)
        self.import_button.grid(row=0, column=4, pady=5, padx=(0, 5), sticky="ew")

        self.check_api_status()

    def check_api_status(self):
        try:
            api_url = f"http://{self.host}:{self.port}/is_api_online"
            response = requests.get(api_url)
            response.raise_for_status()
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
            self.disable_buttons()
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao conectar à API: {e}")
            self.disable_buttons()

    def disable_buttons(self):
        self.add_button.config(state=tk.DISABLED)
        self.view_button.config(state=tk.DISABLED)
        self.delete_button.config(state=tk.DISABLED)
        self.renew_button.config(state=tk.DISABLED)
        self.add_product_button.config(state=tk.DISABLED)
        self.delete_product_button.config(state=tk.DISABLED)
        self.import_button.config(state=tk.DISABLED)

    def add_subscription(self):
        client_name = self.client_name_entry.get().strip()
        selected_product = self.product_listbox.get(tk.ACTIVE)
        end_date_str = self.end_date_entry.get().strip()
        license_key = self.license_key_entry.get().strip() or None

        if not all([client_name, selected_product, end_date_str]):
            self.handle_error("Erro", "Preencha todos os campos, exceto Chave de Licença.")
            return

        try:
            self.check_api_status()

            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            if end_date < datetime.today().date():
                raise ValueError("A data de término não pode estar no passado.")

            api_url = f"http://{self.host}:{self.port}/add_subscription"
            new_subscription = {
                "client_name": client_name,
                "product_name": selected_product,
                "end_date": end_date_str,
                "license_key": license_key
            }
            response = requests.post(api_url, json=new_subscription)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura adicionada com sucesso.")
        except ValueError as e:
            self.handle_error("Erro", str(e))
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao adicionar assinatura: {e}")

    
    def view_subscriptions(self):
        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/view_subscriptions"
            response = requests.get(api_url)
            response.raise_for_status()

            subscriptions = response.json()

            if subscriptions:
                self.view_window = tk.Toplevel(self.master)
                self.view_window.title("Visualizar Assinaturas")

                screen_width = self.view_window.winfo_screenwidth()
                screen_height = self.view_window.winfo_screenheight()
                x = (screen_width - self.view_window.winfo_reqwidth()) // 2
                y = (screen_height - self.view_window.winfo_reqheight()) // 2
                self.view_window.geometry("+{}+{}".format(x, y))

                search_frame = ttk.Frame(self.view_window)
                search_frame.grid(row=0, column=0, columnspan=5, padx=10, pady=5, sticky="ew")

                search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
                search_entry.grid(row=0, column=0, padx=(0, 5), sticky="ew")

                search_button = ttk.Button(search_frame, text="Procurar", command=self.filter_subscriptions)
                search_button.grid(row=0, column=1, padx=(5, 0))

                self.subscriptions_frame = ttk.Frame(self.view_window)
                self.subscriptions_frame.grid(row=1, column=0, columnspan=5, padx=10, pady=5, sticky="ew")

                columns = ("Índice", "Nome do Cliente", "Nome do Produto", "Data de Término", "Chave de Licença")
                self.tree = ttk.Treeview(self.subscriptions_frame, columns=columns, show="headings")
                self.tree.heading("Índice", text="Índice")
                self.tree.heading("Nome do Cliente", text="Nome do Cliente")
                self.tree.heading("Nome do Produto", text="Nome do Produto")
                self.tree.heading("Data de Término", text="Data de Término")
                self.tree.heading("Chave de Licença", text="Chave de Licença")

                self.tree.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)

                # Adding scrollbar
                scrollbar = ttk.Scrollbar(self.subscriptions_frame, orient=tk.VERTICAL, command=self.tree.yview)
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                self.tree.configure(yscroll=scrollbar.set)

                self.render_subscriptions(subscriptions)
            else:
                self.handle_error("Erro", "Nenhuma assinatura encontrada.")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao procurar assinaturas: {e}")


    def render_subscriptions(self, subscriptions):
        # Clear the current contents of the Treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Insert new data into the Treeview
        for index, subscription in enumerate(subscriptions, start=1):
            self.tree.insert("", "end", values=(
                index,
                subscription["client_name"],
                subscription["product_name"],
                subscription["end_date"],
                subscription.get("license_key", "")
            ))



    def filter_subscriptions(self):
        search_term = self.search_var.get().strip().lower()
        api_url = f"http://{self.host}:{self.port}/view_subscriptions"
        response = requests.get(api_url)
        response.raise_for_status()

        subscriptions = response.json()

        filtered_subscriptions = [
            sub for sub in subscriptions
            if search_term in sub["client_name"].lower() or
            search_term in sub["product_name"].lower() or
            (sub.get("license_key") and search_term in sub["license_key"].lower())
        ]

        self.render_subscriptions(filtered_subscriptions)


    def delete_subscription(self):
        try:
            self.check_api_status()

            selected_index = simpledialog.askinteger("Input", "Index da Assinatura:")
            if selected_index is None:
                return

            api_url = f"http://{self.host}:{self.port}/delete_subscription"
            payload = {
                "index": selected_index
            }
            response = requests.delete(api_url, json=payload)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura excluída com sucesso.")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao excluir assinatura: {e}")


    def renew_subscription(self):
        try:
            self.check_api_status()

            client_name = simpledialog.askstring("Input", "Nome do Cliente:")
            if not client_name:
                return

            product_name = simpledialog.askstring("Input", "Nome do Produto:")
            if not product_name:
                return

            additional_days = simpledialog.askinteger("Input", "Dias Adicionais:")
            if additional_days is None:
                return

            api_url = f"http://{self.host}:{self.port}/renew_subscription"
            payload = {
                "client_name": client_name,
                "product_name": product_name,
                "additional_days": additional_days
            }
            response = requests.post(api_url, json=payload)
            response.raise_for_status()
            messagebox.showinfo("Sucesso", "Assinatura renovada com sucesso.")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao renovar assinatura: {e}")

    
    
    def import_from_excel(self):
        def parse_date(date_str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime("%Y-%m-%d")
                except ValueError:
                    continue
            raise ValueError(f"Invalid end date: {date_str}")

        try:
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
            if not file_path:
                return

            wb = load_workbook(filename=file_path)

            for sheet in wb.sheetnames:
                ws = wb[sheet]

                for row in ws.iter_rows(min_row=2, values_only=True):
                    client_name, product_name, license_key, end_date = row[:4]
                    if not client_name or not product_name or not end_date:
                        continue

                    try:
                        # Add product to the API
                        api_url_add_product = f"http://{self.host}:{self.port}/add_product"
                        product_data = {"product_name": product_name}
                        response_add_product = requests.post(api_url_add_product, json=product_data)

                        # Ignore if the product already exists
                        if response_add_product.status_code not in [200, 400]:
                            error_message = response_add_product.json().get("error", "Unknown error")
                            self.handle_error("Error", f"Failed to add product '{product_name}': {error_message}")
                            continue

                        # Parse end date
                        if isinstance(end_date, datetime):
                            end_date_str = end_date.strftime("%Y-%m-%d")
                        else:
                            end_date_str = end_date

                        try:
                            end_date_parsed = parse_date(end_date_str)
                        except ValueError:
                            self.handle_error("Error", f"Invalid end date format for {client_name}: {end_date_str}")
                            continue

                        if datetime.strptime(end_date_parsed, "%Y-%m-%d").date() < datetime.today().date():
                            raise ValueError(f"Invalid end date for {client_name}: {end_date_str}")

                        # Prepare subscription data
                        new_subscription = {
                            "client_name": client_name,
                            "product_name": product_name,
                            "license_key": license_key or None,
                            "end_date": end_date_parsed
                        }

                        # Add subscription to the API
                        api_url_add_subscription = f"http://{self.host}:{self.port}/add_subscription"
                        response_add_subscription = requests.post(api_url_add_subscription, json=new_subscription)

                        # Check for subscription addition errors
                        if response_add_subscription.status_code != 200:
                            error_message = response_add_subscription.json().get("error", "Unknown error")
                            self.handle_error("Error", f"Failed to add subscription for '{client_name}': {error_message}")
                            continue

                    except ValueError as e:
                        self.handle_error("Error", str(e))
                    except requests.RequestException as e:
                        self.handle_error("Error", f"Failed to import subscription: {e}")

            # Refresh the product list
            self.update_product_list()
            messagebox.showinfo("Success", "Subscription import completed.")
        except FileNotFoundError:
            self.handle_error("Error", "File not found.")
        except Exception as e:
            self.handle_error("Error", f"Failed to import subscriptions: {e}")




    def update_product_list(self):
        self.product_listbox.delete(0, tk.END)
        try:
            self.check_api_status()

            api_url = f"http://{self.host}:{self.port}/get_products"
            response = requests.get(api_url)
            response.raise_for_status()

            products = response.json()

            # Debug statement to print the products
            print("Products fetched from API:", products)

            if isinstance(products, list):
                for product in products:
                    if isinstance(product, str):
                        self.product_listbox.insert(tk.END, product)
                    else:
                        print(f"Unexpected product format: {product}")
                        messagebox.showerror("Erro", f"Formato de produto inesperado: {product}")
            else:
                print(f"Unexpected products format: {products}")
                messagebox.showerror("Erro", f"Formato inesperado para produtos: {products}")
        except requests.ConnectionError:
            self.handle_error("Erro", "Falha ao conectar à API: Problema de conexão.")
        except requests.RequestException as e:
            self.handle_error("Erro", f"Falha ao buscar produtos: {e}")
        except Exception as e:
            self.handle_error("Erro", f"Erro inesperado: {e}")


    def add_product(self):
        new_product = self.new_product_entry.get().strip()
        if new_product:
            try:
                self.check_api_status()

                api_url = f"http://{self.host}:{self.port}/add_product"
                response = requests.post(api_url, json={"product_name": new_product})
                response.raise_for_status()

                # Update the product list to include the new product
                self.update_product_list()
                messagebox.showinfo("Sucesso", "Produto adicionado com sucesso.")
            except requests.RequestException as e:
                self.handle_error("Erro", f"Falha ao adicionar produto: {e}")
            except Exception as e:
                self.handle_error("Erro", f"Erro inesperado: {e}")
        else:
            self.handle_error("Erro", "Por favor, insira o nome do novo produto.")


    def delete_product(self):
        selected_product = self.product_listbox.get(tk.ACTIVE)
        if selected_product:
            try:
                self.check_api_status()

                api_url = f"http://{self.host}:{self.port}/delete_product"
                # Send DELETE request with the product name in the request body
                response = requests.delete(api_url, json={"product_name": selected_product})
                response.raise_for_status()
                self.update_product_list()
                messagebox.showinfo("Successo", "Produto apagado com sucesso.")
            except requests.RequestException as e:
                self.handle_error("Erro", f"Falha ao excluir produto: {e}")
        else:
            self.handle_error("Erro", "Por favor, selecione um produto para excluir.")


    def on_product_select(self, event):
        widget = event.widget
        selection = widget.curselection()
        if selection:
            index = selection[0]
            product_name = widget.get(index)

    def create_config_file(self, config_file_path):
        config = configparser.ConfigParser()
        config["Form"] = {
            "host": "localhost",
            "port": "5000"
        }
        with open(config_file_path, "w") as config_file:
            config.write(config_file)

    def handle_error(self, title, message):
        messagebox.showerror(title, message)

if __name__ == "__main__":
    root = tk.Tk()
    app = SubscriptionFormApp(root)
    root.mainloop()